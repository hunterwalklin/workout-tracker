#!/usr/bin/env python3
"""Fetch every tab of the workout Google Sheet and write raw_sheets_data.txt.

No API key required: this downloads the whole spreadsheet as an .xlsx export
(works without auth on a link-shared / "anyone with the link" sheet) and reads
every tab with openpyxl. That gets both the list of tabs AND their contents in
one keyless request.

Output format matches what convert_v2.py expects: a stream of

    === TAB: <tab name> ===
    <tab rows as CSV>

where the tab name is the week's date (e.g. "5/5/25", "5/12"). convert_v2.py
splits on the "=== TAB: ... ===" markers and derives each week's date from the
tab name.

Config comes from the environment:

    SHEET_ID   the spreadsheet id (from the /d/<ID>/edit URL)  [required]

Run: SHEET_ID=... python3 fetch_sheet.py
"""

import csv
import io
import os
import sys
import urllib.request

import openpyxl

OUTPUT_FILE = "raw_sheets_data.txt"


def download_xlsx(sheet_id):
    url = f"https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=xlsx"
    req = urllib.request.Request(url, headers={"User-Agent": "workout-tracker-fetch"})
    with urllib.request.urlopen(req, timeout=60) as resp:
        return resp.read()


def cell_to_str(value):
    """Render a cell value the way the old CSV dump did (empty for blanks)."""
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value)


def rows_to_csv(ws):
    """Render a worksheet as CSV text, quoting every field (matches raw dump)."""
    buf = io.StringIO()
    writer = csv.writer(buf, quoting=csv.QUOTE_ALL, lineterminator="\n")
    for row in ws.iter_rows(values_only=True):
        writer.writerow([cell_to_str(c) for c in row])
    return buf.getvalue()


def main():
    sheet_id = os.environ.get("SHEET_ID", "").strip()
    if not sheet_id:
        sys.exit("SHEET_ID must be set in the environment.")

    data = download_xlsx(sheet_id)
    wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)

    titles = wb.sheetnames
    if not titles:
        sys.exit("No tabs found — check SHEET_ID and that the sheet is link-shared.")

    sections = []
    for title in titles:
        ws = wb[title]
        sections.append(f"=== TAB: {title} ===\n{rows_to_csv(ws)}")

    with open(OUTPUT_FILE, "w") as f:
        f.write("\n".join(sections))

    print(f"Wrote {OUTPUT_FILE}: {len(titles)} tabs ({', '.join(titles)})")


if __name__ == "__main__":
    main()
